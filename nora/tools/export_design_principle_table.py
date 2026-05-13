"""Export the design-principle table to an Excel workbook."""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


ROWS = [
    # (Section header, Item, Short description, Why required)
    ("Required Sections",
     "1. YAML Frontmatter",
     "Machine-readable header declaring name, description, argument-hint, allowed tools, and flags.",
     "Serves as the contract between the skill and the Claude Code harness: enforces tool allowlisting (least privilege), enables orchestrator routing, and exposes opt-in flags without forking the skill."),

    ("Required Sections",
     "2. Overview / Purpose Statement",
     "A 2-5 sentence block explaining what the skill produces, what it consumes, and when to reach for it.",
     "Pipelines are composed by both humans and orchestrator skills; the overview is the only section read at routing time, preventing scope overlap between neighboring skills."),

    ("Required Sections",
     "3. Constants / Configurable Thresholds",
     "A named block of tuneable parameters with defaults (e.g., MAX_ROUNDS, SCORE_THRESHOLD, MAX_PRIMARY_CLAIMS).",
     "Enforces parsimony, guarantees termination of refinement loops, and makes operating envelopes reproducible across colleagues and sessions."),

    ("Required Sections",
     "4. Workflow / Phases",
     "Numbered sequence of phases, each with explicit entry conditions, actions, and exit conditions.",
     "Long-running research workflows need resumability, debuggability, and composability; phase structure localises failures and lets orchestrators skip completed work."),

    ("Required Sections",
     "5. Checkpoint / State Persistence",
     "Explicit state file path, JSON schema, recovery logic, and expiry rules (e.g., REVIEW_STATE.json).",
     "Claude Code sessions are stateless across invocations; without persisted state, every interruption forces a full restart and reviewer/refiner memory is lost."),

    ("Required Sections",
     "6. Canonical Output Paths",
     "A fixed list of every file the skill writes, relative to the project root.",
     "Downstream skills and generate-report depend on stable paths; ad-hoc paths break chaining and cause silent stale-file reads."),

    ("Required Sections",
     "7. Decision Rules / Branching Logic",
     "Explicit if/then rules for forks within the skill (e.g., Track A vs. Track B, code-figure vs. prompt-figure).",
     "Without explicit rules the LLM infers branches from context, causing non-deterministic method choices across sessions."),

    ("Required Sections",
     "8. Guardrails / 'Do NOT' Rules",
     "Explicit list of prohibited behaviours (no fabrication, no self-scoring, no scope creep, no silent failure).",
     "Guardrails are countermeasures against observed LLM failure modes; each rule exists because that exact failure has been seen in testing."),

    ("Required Sections",
     "9. Evidence Discipline",
     "Rules governing how claims are sourced, verified, and traced (APPROVED_CLAIMS.md, claim-to-evidence matrices, placeholder markers).",
     "Load-bearing for research integrity; without it the pipeline produces convincing but unverifiable artifacts, the worst outcome for an academic agent."),

    ("Required Sections",
     "10. Generator-Evaluator Separation",
     "Explicit rules ensuring the entity producing content never evaluates its own quality (Codex MCP reviews, Claude revises).",
     "Self-evaluation causes mode collapse in refinement loops; separation keeps reviewer context uncontaminated and external-scrutiny-ready."),

    ("Required Sections",
     "11. Composability / Pipeline Integration",
     "Documentation of upstream producers, downstream consumers, and standalone vs. pipeline invocation paths.",
     "No skill operates in isolation; undocumented composability turns skills into black boxes that only work in one exact sequence."),

    ("Required Sections",
     "12. Key Rules Summary",
     "A compressed bullet-point summary of the most critical rules, always including the large-file Bash-heredoc fallback.",
     "Full SKILL.md files run to thousands of tokens; the summary acts as recency-biased reinforcement so critical rules survive long-context attention decay."),

    ("Required Sections",
     "13. Audit Trail (PROJ_NOTES.md Logging)",
     "Every skill appends a dated one-line entry to output/PROJ_NOTES.md describing what was produced.",
     "Provides provenance, progress tracking across multi-day campaigns, and pipeline debugging without archaeological artifact hunting."),

    ("Required Sections",
     "14. Harness Engineering",
     "Non-LLM scaffolding: lifecycle hooks (pre_tool_use, post_tool_use, stop_hook, notification), persistent state (handoff.json, memory/MEMORY.md, REVIEW_STATE.json), and declarative config (settings.json, .mcp.json, CLAUDE.md, configs/default.yaml).",
     "Long-horizon autonomous research exceeds any single context window; the harness externalises permissioning, logging, checkpointing, and recovery from skill logic, defeats 'context anxiety', and makes multi-session runs deterministically resumable. Skills encode intent; the harness encodes guarantees."),

    ("Advanced Optimisations",
     "A. Pre/Post-Condition Contracts",
     "Machine-checkable YAML contracts declaring each phase's preconditions, postconditions, and on_failure action.",
     "Prose conditions drift under interpretation; formal contracts let skills self-validate and let orchestrators pre-flight check downstream invocations, eliminating malformed-input bugs."),

    ("Advanced Optimisations",
     "B. Artifact Versioning and Immutability",
     "Append-only versioned artifacts (FINAL_PROPOSAL_v1.md, v2.md, ...) with a symlink to the current version.",
     "Overwriting destroys refinement history; immutable versioning enables rollback, reviewer traceability ('why did v3 drop this?'), and counterfactual comparison."),

    ("Advanced Optimisations",
     "C. Confidence-Gated Progression",
     "Three-tier pass/flag/halt gate (High / Medium / Low) instead of a binary score threshold.",
     "Binary gating forces a trade-off between permissiveness and strictness; the medium tier lets work continue while flagging uncertain decisions for asynchronous human review."),

    ("Advanced Optimisations",
     "D. Semantic Dependency Graph",
     "A machine-readable DEPENDENCY_GRAPH.yaml mapping each skill's reads, writes, requires_before, and enables_after.",
     "Enables automated precondition checks, parallel execution of independent branches, impact analysis on output-format changes, and DAG visualisation for onboarding."),

    ("Advanced Optimisations",
     "E. Structured Error Taxonomy",
     "Standard error classes (MISSING_INPUT, TOOL_UNAVAILABLE, QUALITY_BELOW_THRESHOLD, STATE_CORRUPTION, RESOURCE_LIMIT, EXTERNAL_TIMEOUT) with prescribed responses.",
     "Consistent error handling across skills eliminates the silent-degradation class of bugs; every failure mode has a known, documented response."),

    ("Advanced Optimisations",
     "F. Skill Telemetry and Cost Accounting",
     "Per-run JSON telemetry: rounds used, external-LLM calls, input/output tokens, wall-clock minutes, final score, artifacts produced.",
     "Without cost accounting users cannot distinguish efficient convergence from budget burn; telemetry drives data-driven threshold tuning and reviewer-backend selection."),

    ("Advanced Optimisations",
     "G. Degradation-Aware Fallback Chains",
     "Explicit fallback chains annotated with quality tiers (Codex MCP HIGH -> Claude subagent MEDIUM -> self-review LOW).",
     "The current graceful-degradation pattern silently lowers quality; quality annotations let downstream skills adjust trust (e.g., add an extra review round after a MEDIUM upstream review)."),

    ("Advanced Optimisations",
     "H. Prompt Modularisation and Slot Filling",
     "Reusable prompt templates with named slots ({{ARTIFACT_TYPE}}, {{VENUE}}, {{CRITERIA_BLOCK}}) stored under templates/prompts/.",
     "Inline prompts are hard to A/B test, reuse, or version; modular templates let prompts evolve independently of skill logic."),

    ("Advanced Optimisations",
     "I. Multi-Perspective Review Ensembles",
     "Parallel reviewer personas per round (ML, GIScience, methods-skeptic) with weighted aggregation and conflict flagging.",
     "Single-perspective reviews have disciplinary blind spots; ensembles surface cross-domain issues early, trading 3x review tokens for fewer wasted revision rounds."),

    ("Advanced Optimisations",
     "J. Skill Self-Test Suites",
     "A tests/ directory per skill with synthetic scenarios (resume from checkpoint, max-rounds halt, missing input, large-file fallback).",
     "Skills are complex natural-language programs; markdown test cases enable regression testing after edits and document expected behaviour at boundary conditions."),
]


def write_excel(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Design Principles"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")

    section_fill_required = PatternFill("solid", fgColor="DBEAFE")
    section_fill_advanced = PatternFill("solid", fgColor="FEF3C7")

    thin = Side(border_style="thin", color="9CA3AF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Section", "Item", "Short Description", "Why Required"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for section, item, desc, why in ROWS:
        ws.append([section, item, desc, why])
        row = ws.max_row
        fill = section_fill_required if section == "Required Sections" else section_fill_advanced
        for col_idx in range(1, 5):
            cell = ws.cell(row=row, column=col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if col_idx == 1:
                cell.fill = fill
                cell.font = Font(bold=True, name="Calibri")
            else:
                cell.font = Font(name="Calibri")

    widths = {1: 22, 2: 38, 3: 60, 4: 75}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 26
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 72

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


if __name__ == "__main__":
    out = Path(__file__).resolve().parent.parent / "output" / "reports" / "DESIGN_PRINCIPLES_TABLE.xlsx"
    write_excel(out)
    print(f"Wrote {out}")
