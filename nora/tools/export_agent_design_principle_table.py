"""Export the agent design-principle table to an Excel workbook."""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


ROWS = [
    # (Section header, Item, Short description, Why required)
    ("Required Sections",
     "1. YAML Frontmatter",
     "Machine-readable header declaring name, description ('Use this agent to' bullets), and an explicit tool allowlist (never 'all' except orchestrator).",
     "Enables parent skills and the orchestrator to route the right subtask to the right agent and enforces the tightest tool allowlist in the system, because agents run unattended in a separate context with no user gate during execution."),

    ("Required Sections",
     "2. Persona and Role Statement",
     "A 2-4 sentence block declaring who the agent is and what scope it owns.",
     "Agents start in a fresh sub-context with no conversational history; the persona is the first thing the sub-context reads, seeds role-consistent behaviour, and disambiguates scope overlap between neighbouring agents."),

    ("Required Sections",
     "3. Context Isolation Contract",
     "Explicit statement of what the agent reads from disk on startup and what it does NOT assume from the caller's context.",
     "The parent's context is invisible to the agent, so everything the agent needs must be fetchable from files or passed explicitly as $ARGUMENTS; this is the architectural reason NORA is file-driven."),

    ("Required Sections",
     "4. Single-Responsibility Scope",
     "A declaration that the agent performs exactly one discrete subtask, not a pipeline.",
     "Composite agents collapse specialists into one context, erase parallelism, and defeat generator-evaluator separation because the same context that produces cannot objectively critique its own output."),

    ("Required Sections",
     "5. Narrow Tool Allowlist (Least Privilege)",
     "An explicit tools list containing only what the agent needs (e.g., peer-reviewer has Read+Write only).",
     "Because agents run unattended, the blast radius of a single call is bounded only by the allowlist; reviewer agents that can Write anywhere can corrupt the artifact they review."),

    ("Required Sections",
     "6. Deterministic I/O Contract",
     "Specification of required $ARGUMENTS, files consulted on disk, and the structured output shape returned to the caller.",
     "Callers must predict the return payload well enough to parse it and feed it downstream without an LLM re-read, which would waste tokens and reintroduce context pollution."),

    ("Required Sections",
     "7. Evaluator-vs-Producer Role Lock",
     "A rule declaring whether the agent produces content or evaluates content — never both.",
     "Agent-side enforcement of the generator-evaluator separation principle; prevents refinement loops from converging on outputs that score well against their own priors but fail external scrutiny."),

    ("Required Sections",
     "8. Persona-Grounded Expertise",
     "An embedded block of domain knowledge (CRS rules, APA 7 rules, spatial-method decision rules, venue-formatting deltas) the agent consults on every invocation.",
     "Agents start cold; baked-in priors compensate for the absence of conversational history and keep output consistent with discipline-specific conventions across runs."),

    ("Required Sections",
     "9. Structured Output Format (Return Schema)",
     "A concrete Markdown/JSON template with named fields and example values (e.g., 'Score: X.X (N:X, R:X, L:X, C:X, I:X)').",
     "Gives the agent a self-checklist, gives the caller a parsing anchor, and gives auditors a documented expectation; prevents format drift between invocations."),

    ("Required Sections",
     "10. Cold-Read Discipline (Evaluator Agents Only)",
     "Explicit instruction for reviewer agents to evaluate the artifact as if encountering it for the first time, without author rationale.",
     "The reviewer sub-context's value is precisely that it has no memory of how the draft was produced; without this, the review inherits the author's framing and misses blind spots."),

    ("Required Sections",
     "11. Cost and Parallelism Awareness",
     "A note on whether the agent can be invoked in parallel with others and an estimate of its typical token footprint.",
     "Each agent call spawns an independently billed sub-context; the orchestrator needs this to parallelise independent calls and serialise ones with data dependencies."),

    ("Required Sections",
     "12. Error and Degradation Handling",
     "An explicit list of failure modes and the agent's prescribed deterministic response (retries, fallbacks, escalation rubrics).",
     "Agents cannot ask the user for help mid-run; every foreseeable failure must have a documented response or the agent halts silently with no diagnostic trail."),

    ("Required Sections",
     "13. Audit Contribution",
     "Agent appends to output/PROJ_NOTES.md itself or returns enough metadata for the parent skill to log on its behalf.",
     "Extends the skill-level audit trail across the skill-agent boundary; without it, artifacts arrive with no record of which agent in which invocation produced them."),

    ("Advanced Optimisations",
     "A. Agent Contract Tests",
     "A .claude/agents/tests/<agent>/ directory with synthetic input fixtures and expected-output schemas.",
     "Catches prompt regressions (an edit that silently changes the return schema) before they reach the pipeline and documents expected behaviour at boundary conditions."),

    ("Advanced Optimisations",
     "B. Parallel Invocation Manifest",
     "Each agent declares its parallelism class and max_concurrent in frontmatter (independent / depends_on / global_singleton).",
     "The orchestrator can validate a parallel dispatch plan statically, catching dependency violations before invocation instead of after a failed run."),

    ("Advanced Optimisations",
     "C. Return-Payload JSON Schema",
     "A JSON Schema shipped alongside each agent under .claude/agents/schemas/.",
     "Parent skills can validate agent output programmatically, rejecting malformed returns with a clear diagnostic instead of silently consuming a partial payload."),

    ("Advanced Optimisations",
     "D. Persona Composition over Inheritance",
     "Shared domain knowledge extracted into reusable blocks under .claude/agents/_shared/ (e.g., geo_conventions.md, apa7_rules.md).",
     "Rules duplicated across three agents drift apart over time; shared blocks keep rules synchronised and make a single edit propagate to every consumer."),

    ("Advanced Optimisations",
     "E. Cold-Read Assertion",
     "Orchestrator strips author rationale from the payload before invoking a reviewer; the reviewer asserts author_context_present == false before proceeding.",
     "Makes the cold-read property structurally enforced rather than just prompted; a prompt rule can be overridden by a chatty caller, a payload check cannot."),

    ("Advanced Optimisations",
     "F. Token-Budget Declaration",
     "Each agent declares expected input/output token ranges in frontmatter; orchestrator refuses invocation when the payload exceeds the upper bound.",
     "Prevents runaway dispatches that would time out or blow the context window and makes per-agent cost legible in pre-flight planning."),

    ("Advanced Optimisations",
     "G. Reviewer Ensemble for High-Stakes Gates",
     "Dispatch three independent peer-reviewer invocations in parallel sub-contexts for final-acceptance gates; aggregate the verdicts.",
     "Three independent sub-contexts cannot influence each other, eliminating the intra-context consensus bias of a single invocation simulating three reviewers."),

    ("Advanced Optimisations",
     "H. Reviewer Memory (Adversarial Continuity)",
     "Reviewer agents explicitly read and append to memory/REVIEWER_MEMORY.md so suspicions raised in round N survive into round N+1.",
     "Extends the auto-review-loop's adversarial continuity to local Claude-subagent reviewers, closing a quality gap in the fallback path when Codex MCP is unavailable."),

    ("Advanced Optimisations",
     "I. Capability Probe Instead of Static Allowlists",
     "Agents declare capability tags (retrieves_papers, formats_citations, reviews_spatial_methods) and orchestrators select by capability, not by name.",
     "Decouples orchestrators from agent names, enabling drop-in replacement when a better agent exists (e.g., literature-scout swapped for semantic-scout-v2) without caller code changes."),

    ("Advanced Optimisations",
     "J. Agent Versioning and Deprecation",
     "Agents carry a version field and support side-by-side coexistence (peer-reviewer@1, peer-reviewer@2); orchestrator pins a version per pipeline run.",
     "Prompt edits to a reviewer agent silently change every downstream score; versioning makes prompt evolution auditable and lets experimental prompts ship without disturbing production."),
]


def write_excel(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Agent Design Principles"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")

    section_fill_required = PatternFill("solid", fgColor="E0E7FF")
    section_fill_advanced = PatternFill("solid", fgColor="FCE7F3")

    thin = Side(border_style="thin", color="9CA3AF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Section", "Item", "Short Description", "Why Required"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for section, item, desc, why in ROWS:
        ws.append([section, item, desc, why])
        row = ws.max_row
        fill = section_fill_required if section == "Required Sections" else section_fill_advanced
        for col_idx in range(1, 5):
            cell = ws.cell(row=row, column=col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if col_idx == 1:
                cell.fill = fill
                cell.font = Font(bold=True, name="Calibri")
            else:
                cell.font = Font(name="Calibri")

    widths = {1: 22, 2: 44, 3: 62, 4: 78}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 26
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 78

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


if __name__ == "__main__":
    out = Path(__file__).resolve().parent.parent / "output" / "reports" / "AGENT_DESIGN_PRINCIPLES_TABLE.xlsx"
    write_excel(out)
    print(f"Wrote {out}")
